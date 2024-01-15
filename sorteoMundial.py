import tkinter as tk
from tkinter import ttk
import random
import openpyxl
from openpyxl.styles import Font
import subprocess

#Presentación inicial

grupo_x = ''

def delete_window_i():
    ventana.destroy()

ventana = tk.Tk()
ventana.geometry("400x300")
ventana.title("Sorteo Mundial Sub 13 Qatar 2023")

imagen = tk.PhotoImage(file="banderas/fifa.png").subsample(5)
label = ttk.Label(image = imagen)
label.pack(side = "top", pady= 80)

boton = tk.Button(ventana, text = "Siguiente", command = delete_window_i)
boton.pack(side ="bottom")

ventana.mainloop()

#Función de ventanas

def ventanas(pais):
    def delete_window():
        ventana.destroy()
    ventana = tk.Tk()
    ventana.geometry("600x500")
    ventana.title("Sorteo Mundial Sub 13 Qatar 2023")
    imagen = tk.PhotoImage(file=f"banderas/{pais}.png")
    if imagen.width() > 2000:
        imagen = imagen.subsample(8)
    elif imagen.width() > 1000:
        imagen = imagen.subsample(4)
    elif imagen.width() > 800:
        imagen = imagen.subsample(3)
    elif imagen.width() > 500:
        imagen = imagen.subsample(2)

    label = ttk.Label(image = imagen)
    label.pack(side = "top", pady= 50)

    etiqueta = tk.Label(ventana, text = f"{pais}", font= ("Calibri",20))
    etiqueta.pack(side = "top", pady= 20)

    boton = tk.Button(ventana, text = "Siguiente", command= delete_window)
    boton.pack(side ="bottom")

    ventana.mainloop()

def ventanas2(posicion):
    def delete_window():
        ventana.destroy()
    ventana = tk.Tk()
    ventana.geometry("600x400")
    ventana.title("Sorteo Mundial Sub 13 Qatar 2023")

    label = ttk.Label(text = f"{posicion}", font = ("Calibri",120))
    label.pack(side = "top", pady= 50)

    boton = tk.Button(ventana, text = "Siguiente", command= delete_window)
    boton.pack(side ="bottom")

    ventana.mainloop()

def ventanas3():
    def contestar():
        global grupo_x
        texto = entrada.get()
        grupo_x = texto
        ventana.destroy()
    
    ventana = tk.Tk()
    ventana.geometry("600x500")
    ventana.title("Sorteo Mundial Sub 13 Qatar 2023")

    label = ttk.Label(text = f"¿En qué grupo desea continuar", font = ("Calibri",20))
    label.pack(side = "top", pady= 50)

    entrada = tk.Entry(ventana, font = ("Calibri", 20))
    entrada.pack(side = "top", pady = 20)

    boton = tk.Button(ventana, text = "Siguiente", command= contestar)
    boton.pack(side ="bottom")

    ventana.mainloop()
    return grupo_x


# Define los equipos en cada bombo
bombo1 = ["Qatar", "Brasil", "Colombia","Japón", "Estados Unidos", "Argentina"]
bombo2 = ["Zambia", "Nueva Zelanda", "España", "Italia", "México", "Inglaterra"]
bombo3 = ["Ecuador", "Senegal", "Corea del Sur", "Marruecos", "Uzbekistán", "Croacia"]
bombo4 = ["Canadá", "Portugal", "Panamá", "Emiratos Árabes Unidos", "Sierra Leona", "Fiyi"]

A = ["A1","A2","A3","A4"]
B = ["B1","B2","B3","B4"]
C = ["C1","C2","C3","C4"]
D = ["D1","D2","D3","D4"]
E = ["E1","E2","E3","E4"]
F = ["F1","F2","F3","F4"]

#Abre libro de Excel

ruta = "Torneos.xlsx"
archivo = openpyxl.load_workbook(ruta)

if "Mundial" in archivo.sheetnames:
    hoja = archivo["Mundial"]
else:
    hoja = archivo.create_sheet("Mundial")

# Define los grupos
grupos = {}
for i in range(1, 7):
    grupos[f"Grupo {chr(64+i)}"] = ['','','','']
    celda = hoja[f'{chr(64+i+9)}1']
    celda.value = f'Grupo {chr(64+i)}'
    celda.font = Font(bold=True)
    # Guardar el libro de Excel
    archivo.save("Torneos.xlsx")
    #subprocess.run(f'start "" "Torneos.xlsx"', shell=True)


print(grupos)

#Realizamos el sorteo del primer bombo

#A1 Anfitrión
#Escoge
equipo = bombo1[0]
posicion = A[0]
#Visualiza el pais
ventanas(equipo)
#Coloca en los grupos
grupos[f"Grupo {chr(65)}"][int(posicion[1])-1] = equipo
#Adjunta en la celda de excel
celda = hoja.cell(row=2, column=10)
celda.value = equipo
#Guarda
archivo.save("Torneos.xlsx")
#visualiza la posicion
ventanas2(posicion)
#Borra la balota
bombo1.remove(equipo)
A.remove(posicion)

letras = [A,B,C,D,E,F]

print(grupos)

for i in range(5):
    #Escoge
    equipo = random.choice(bombo1)
    posicion = letras[i+1][0]
    #Visualiza el pais
    ventanas(equipo)
    #Coloca en los grupos
    grupos[f"Grupo {chr(66 + i)}"][int(posicion[1])-1] = equipo
    #Adjunta en la celda de excel
    celda = hoja.cell(row=2, column=i+11)
    celda.value = equipo
    #Guarda
    archivo.save("Torneos.xlsx")
    #visualiza la posicion
    ventanas2(posicion)
    #Borra la balota
    bombo1.remove(equipo)
    letras[i+1].remove(posicion)
    print(grupos)



bombos = [bombo2,bombo3,bombo4]

#Realizamos el sorteo
for i in range(3):
    for j in range(6):
        #Escoge
        equipo = random.choice(bombos[i])
        #Visualiza el pais
        ventanas(equipo)
        ventanas3()
        grupo_x = ord(grupo_x)
        posicion = random.choice(letras[grupo_x-65])
        #Coloca en los grupos
        grupos[f"Grupo {chr(grupo_x)}"][int(posicion[1])-1] = equipo
        #Adjunta en la celda de excel
        celda = hoja.cell(row=int(posicion[1])+1, column=grupo_x - 65 +10)
        celda.value = equipo
        #Guarda
        archivo.save("Torneos.xlsx")
        #Visualiza la posición
        ventanas2(posicion)
        #Borra la balota
        bombos[i].remove(equipo)
        letras[grupo_x-65].remove(posicion)
        print(grupos)

subprocess.run(f'start "" "Torneos.xlsx"', shell=True)

""" #try:
ruta = "Torneos 2.xlsx"
archivo = openpyxl.load_workbook(ruta)

if "Concacaf México" in archivo.sheetnames:
    hoja = archivo["Concacaf México"]
else:
    hoja = archivo.create_sheet("Concacaf México")

for grupo, paises in grupos.items():
    celda = hoja[f'{chr(8 + ord(grupo[6:7]))}8']
    celda.value = f'Grupo {grupo[6:7]}'
    celda.font = Font(bold=True)
    columna = ord(grupo[6:7]) - ord('A') + 9
    fila = 9
    for pais in paises:
        celda = hoja.cell(row=fila, column=columna)
        celda.value = pais
        fila += 1

# Guardar el libro de Excel
archivo.save("Torneos 2.xlsx")
subprocess.run(f'start "" "Torneos 2.xlsx"', shell=True) """