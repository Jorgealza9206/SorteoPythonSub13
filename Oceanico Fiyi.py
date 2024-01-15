import random
import openpyxl
from openpyxl.styles import Font
import subprocess

# Define los equipos en cada bombo
bombo1 = ["Fiyi", "Nueva Zelanda", "Papúa Nueva Guinea"]
bombo2 = ["Tahití", "Nueva Caledonia", "Vanuatu"]
bombo3 = ["Islas Salomón", "Samoa", "Tonga"]
bombo4 = ["Samoa Estadounidense", "Islas Cook",""]

# Define los grupos
grupos = {}
for i in range(1, 4):
    grupos[f"Grupo {chr(64+i)}"] = []

#print(grupos)

#Realizamos el sorteo del primer bombo

#A1 Anfitrión
equipo = bombo1[0]
grupos[f"Grupo {chr(65)}"].append(equipo)
bombo1.remove(equipo)

for i in range(2):
    equipo = random.choice(bombo1)
    grupos[f"Grupo {chr(66 + i)}"].append(equipo)
    bombo1.remove(equipo)

#print(grupos)

bombos = [bombo2,bombo3,bombo4]

#Realizamos el sorteo
for i in range(3):
    for j in range(3):
        equipo = random.choice(bombos[i])
        grupos[f"Grupo {chr(65 + j)}"].append(equipo)
        bombos[i].remove(equipo)

print(grupos)

#try:
ruta = "Torneos 2.xlsx"
archivo = openpyxl.load_workbook(ruta)

if "Oceanico Fiyi" in archivo.sheetnames:
    hoja = archivo["Oceanico Fiyi"]
else:
    hoja = archivo.create_sheet("Oceanico Fiyi")


fila = 1
columna = 0

for grupo, paises in grupos.items():
    hoja[f'{chr(65 + columna)}{fila}'] = f"Grupo {chr(65 + columna)}"
    hoja[f'{chr(65 + columna)}{fila}'].font = Font(bold=True)
    for pais in paises:
        fila += 1
        hoja[f'{chr(65 + columna)}{fila}'] = f"{pais}"
        #hoja.write(fila, columna, pais)
    columna += 1
    fila = 1

archivo.save(ruta)
subprocess.run(f'start "" "Torneos 2.xlsx"', shell=True)