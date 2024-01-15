import random
import openpyxl
from openpyxl.styles import Font
import subprocess

# Define los equipos en cada bombo
bombo1 = ["México", "Estados Unidos", "Honduras"]
bombo2 = ["Panamá", "Costa Rica", "Jamaica"]
bombo3 = ["Trinidad y Tobago", "Guatemala", "Canadá"]
bombo4 = ["El Salvador", "Aruba", "Martinica"]

# Define los grupos
grupos = {}
for i in range(1, 4):
    grupos[f"Grupo {chr(64+i)}"] = []

#print(grupos)

#Realizamos el sorteo del primer bombo

#A1 Anfitrión
equipo = bombo1[0]
grupos[f"Grupo {chr(65)}"].append(equipo)
bombo1.remove(equipo)

for i in range(2):
    equipo = random.choice(bombo1)
    grupos[f"Grupo {chr(66 + i)}"].append(equipo)
    bombo1.remove(equipo)

#print(grupos)

bombos = [bombo2,bombo3,bombo4]

#Realizamos el sorteo
for i in range(3):
    for j in range(3):
        equipo = random.choice(bombos[i])
        grupos[f"Grupo {chr(65 + j)}"].append(equipo)
        bombos[i].remove(equipo)

print(grupos)

#try:
ruta = "Torneos 2.xlsx"
archivo = openpyxl.load_workbook(ruta)

if "Concacaf México" in archivo.sheetnames:
    hoja = archivo["Concacaf México"]
else:
    hoja = archivo.create_sheet("Concacaf México")

for grupo, paises in grupos.items():
    celda = hoja[f'{chr(8 + ord(grupo[6:7]))}8']
    celda.value = f'Grupo {grupo[6:7]}'
    celda.font = Font(bold=True)
    columna = ord(grupo[6:7]) - ord('A') + 9
    fila = 9
    for pais in paises:
        celda = hoja.cell(row=fila, column=columna)
        celda.value = pais
        fila += 1

# Guardar el libro de Excel
archivo.save("Torneos 2.xlsx")
subprocess.run(f'start "" "Torneos 2.xlsx"', shell=True)