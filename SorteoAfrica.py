import random
import openpyxl

# Define los equipos en cada bombo
bombo1 = ["Costa de Marfil", "Sudáfrica", "Senegal"]
bombo2 = ["Zambia", "Ghana", "Nigeria"]
bombo3 = ["Egipto", "Gabón", "Marruecos"]
bombo4 = ["Sierra Leona", "Tanzania", "Kenia"]

# Define los grupos
grupos = {}
for i in range(1, 4):
    grupos[f"Grupo {chr(64+i)}"] = []

#print(grupos)

#Realizamos el sorteo del primer bombo

#A1 Anfitrión
equipo = bombo1[0]
grupos[f"Grupo {chr(65)}"].append(equipo)
bombo1.remove(equipo)

for i in range(2):
    equipo = random.choice(bombo1)
    grupos[f"Grupo {chr(66 + i)}"].append(equipo)
    bombo1.remove(equipo)

#print(grupos)

bombos = [bombo2,bombo3,bombo4]

#Realizamos el sorteo
for i in range(3):
    for j in range(3):
        equipo = random.choice(bombos[i])
        grupos[f"Grupo {chr(65 + j)}"].append(equipo)
        bombos[i].remove(equipo)

print(grupos)

#try:
ruta = "Torneos 2.xlsx"
archivo = openpyxl.load_workbook(ruta)

if "Africano Costa de Marfil" in archivo.sheetnames:
    hoja = archivo["Africano Costa de Marfil"]
else:
    hoja = archivo.create_sheet("Africano Costa de Marfil")


fila = 1
columna = 0

for grupo, paises in grupos.items():
    hoja[f'{chr(65 + columna)}{fila}'] = f"Grupo {chr(65 + columna)}"
    for pais in paises:
        fila += 1
        hoja[f'{chr(65 + columna)}{fila}'] = f"{pais}"
        #hoja.write(fila, columna, pais)
    columna += 1
    fila = 1

archivo.save(ruta)

#except:
#print("Descarga la librería de excel")