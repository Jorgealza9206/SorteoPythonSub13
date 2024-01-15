import random
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Font
import subprocess

bombo1 = ["Guatemala", "Canadá", "Trinidad y Tobago", "Cuba", "El Salvador", "Nicaragua"]
bombo2 = ["Haití", "Puerto Rico", "Surinam", "Bermudas", "Curazao", "República Dominicana"]
bombo3 = ["Granada", "Martinica", "Aruba", "San Cristóbal y Nieves", "Antigua y Barbuda", "Guyana"]
bombo4 = ["Dominica", "Belice", "Islas Virgenes Británicas", "Saint-Martin", "Bahamas", "Barbados"]
bombo5 = ["Anguila", "Islas Turcas y Caicos", "Islas Caimán"]

grupos = {
    'A': [],
    'B': [],
    'C': [],
    'D': [],
    'E': [],
    'F': [],
}

random.shuffle(bombo1)
for i, equipo in enumerate(bombo1):
    grupo = chr(ord('A') + i)
    grupos[grupo].append(equipo)

random.shuffle(bombo2)
for i, equipo in enumerate(bombo2):
    grupo = chr(ord('F') - i)
    grupos[grupo].append(equipo)

random.shuffle(bombo3)
for i, equipo in enumerate(bombo3):
    grupo = chr(ord('A') + i)
    grupos[grupo].append(equipo)

random.shuffle(bombo4)
for i, equipo in enumerate(bombo4):
    grupo = chr(ord('F') - i)
    grupos[grupo].append(equipo)

random.shuffle(bombo5)
for i, equipo in enumerate(bombo5):
    grupo = chr(ord('A') + i)
    grupos[grupo].append(equipo)

for grupo, lista in grupos.items():
    print("Grupo " + grupo)
    for i, pais in enumerate(lista, start=1):
        print(f'{i}. {pais}')

# Crear un nuevo libro de Excel
ruta = "Torneos 2.xlsx"
archivo = load_workbook(ruta)

# Seleccionar la hoja activa
if "Concacaf México" in archivo.sheetnames:
    hoja = archivo["Concacaf México"]
else:
    hoja = archivo.create_sheet("Concacaf México")

for grupo, paises in grupos.items():
    celda = hoja[f'{chr(8 + ord(grupo))}1']
    celda.value = f'Grupo {grupo}'
    celda.font = Font(bold=True)
    columna = ord(grupo) - ord('A') + 9
    fila = 2
    for pais in paises:
        celda = hoja.cell(row=fila, column=columna)
        celda.value = pais
        fila += 1

# Guardar el libro de Excel
archivo.save("Torneos 2.xlsx")
subprocess.run(f'start "" "Torneos 2.xlsx"', shell=True)